import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time
import sys

# Try curl_cffi first (best Cloudflare bypass), fall back to requests
try:
    from curl_cffi import requests
    USE_CURL = True
    IMPERSONATE = "chrome124"
except ImportError:
    import requests
    USE_CURL = False

# ============================================================
#  EDITABLE PARAMETERS
# ============================================================
TIMEFRAME   = "24h"    # Options: 1m, 5m, 1h, 6h, 24h
TOP_N       = 450     # How many tokens to fetch
MIN_VOLUME  = 25000       # Minimum volume filter (0 = no filter)
CHAIN       = "base"  # Chain: base, eth, sol, bsc, etc.

# Filters (remove any you don't want)
FILTERS = ["not_honeypot", "verified", "renounced"]

# ============================================================
#  COOKIE STRING
# ============================================================
#
#  HOW TO GET THE FULL COOKIE STRING (important — get ALL cookies):
#  1. Open https://gmgn.ai in Chrome, let it load fully
#  2. Press F12 -> Network tab -> tick "Preserve log"
#  3. Refresh the page
#  4. In the filter box type:  rank
#  5. Click the first result (e.g. swaps/1h?...)
#  6. Click "Headers" panel -> scroll to "Request Headers"
#  7. Right-click the "cookie" row -> "Copy value"
#     (The full string starts with _ga= or similar)
#  8. Paste the entire string below between the quotes
#
#  NOTE: cf_clearance is tied to your IP + browser User-Agent.
#        The script uses curl_cffi to match Chrome's TLS fingerprint.
#        Install it once:  pip install curl_cffi
#
COOKIE_STRING = "cf_clearance=y0t7bN3DHm_uSZvJYEhiGSO4qiYJO7lyncXsgLP2kiM-1759727174-1.2.1.1-0N_a.l4CDl5.b9rx9BZEFZI5RPYQRAYNeFghFVVTKNyq4AxmQjMYMv0vjcAblxPpj224KN_FllJjfF.394YoeSpRW5NHlSQe6QbIhlkBkTIiAIlUa_CISF4UBrEA7drSYLqprwXvEw3yWDQg2pZHLeMzE65pdTpdX6alXC5rDUS3HsPPLqmF3uLRQKbPGO_MY1U_6izSPKzQo_AlbsASUeKN_xyO1hkroPu6RNYF_74; sid=gmgn%7C6980259ed846f790981e26482f6392ed; _ga_UGLVBMV4Z0=GS1.2.1773746309961506.95068faaf4c1f302c7b5213087f5cdf2.Etbr2vSk0537EhgLeseXpA%3D%3D.mk9%2BmcOlav5i0PKX6DQG8A%3D%3D.khYZicHOkCz1rkthOC%2B7Hw%3D%3D.fSzGp532apqR0Prn7%2B%2BAeg%3D%3D; __cf_bm=BYj8f79i5Pqg8hSlr7TnVaU8iqC30cITFwa.rSyTTsw-1773746852.335027-1.0.1.1-XKCiZbzz6O.pkeD6DL0kpx5ehgfZar6RDYZhI1aYGZ_G3jVCyPXB7NGuChwKmhlRk7CI8Ynd_jzh0WZparyny0Agu7yoYmFikWCADA87SvdTQ8_1V15dAnYqDlWcZ1VG"

# ============================================================
#  SESSION PARAMS  — update from DevTools URL if needed
# ============================================================
DEVICE_ID  = "11f6dfb1-0336-46c4-8e1d-af367e46824e"
FP_DID     = "3193ac961ff03f6fedd3dbc953744c51"
CLIENT_ID  = "gmgn_web_20260317-11797-35b8ac9"
APP_VER    = "20260317-11797-35b8ac9"
TZ_NAME    = "Africa/Lagos"
TZ_OFFSET  = "3600"
# ============================================================


BASE_URL = "https://gmgn.ai/api/v1/rank/{chain}/swaps/{timeframe}"

# Headers used when curl_cffi is NOT available
HEADERS = {
    "User-Agent":       "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept":           "application/json, text/plain, */*",
    "Accept-Language":  "en-US,en;q=0.9",
    "Accept-Encoding":  "gzip, deflate, br",
    "Referer":          "https://gmgn.ai/",
    "Origin":           "https://gmgn.ai",
    "Connection":       "keep-alive",
    "Sec-Fetch-Dest":   "empty",
    "Sec-Fetch-Mode":   "cors",
    "Sec-Fetch-Site":   "same-origin",
    "sec-ch-ua":        '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "Cookie":           COOKIE_STRING,
}


def fetch_tokens():
    url       = BASE_URL.format(chain=CHAIN, timeframe=TIMEFRAME)
    tokens    = []
    page_size = 50
    offset    = 0

    print(f"\n{'='*60}")
    print(f" GMGN Token Scraper  |  Chain: {CHAIN.upper()}  |  Timeframe: {TIMEFRAME}")
    print(f" Target: top {TOP_N} tokens")
    mode = "curl_cffi (Chrome TLS)" if USE_CURL else "requests — run: pip install curl_cffi"
    print(f" Mode   : {mode}")
    print(f"{'='*60}\n")

    while len(tokens) < TOP_N:
        need  = TOP_N - len(tokens)
        limit = min(page_size, need)

        static = (
            f"device_id={DEVICE_ID}"
            f"&fp_did={FP_DID}"
            f"&client_id={CLIENT_ID}"
            f"&from_app=gmgn"
            f"&app_ver={APP_VER}"
            f"&tz_name={TZ_NAME.replace('/', '%2F')}"
            f"&tz_offset={TZ_OFFSET}"
            f"&app_lang=en-US"
            f"&os=web"
            f"&worker=0"
            f"&orderby=volume"
            f"&direction=desc"
            f"&offset={offset}"
            f"&limit={limit}"
        )
        if MIN_VOLUME > 0:
            static += f"&min_volume={MIN_VOLUME}"

        filters_qs = "&".join(f"filters[]={f}" for f in FILTERS)
        full_url   = f"{url}?{static}&{filters_qs}"

        print(f"  Fetching offset={offset}, limit={limit} ...")

        try:
            if USE_CURL:
                resp = requests.get(
                    full_url,
                    headers=HEADERS,
                    impersonate=IMPERSONATE,
                    timeout=20,
                )
            else:
                resp = requests.get(full_url, headers=HEADERS, timeout=20)

            if resp.status_code == 403:
                print("\n  [403 Forbidden] — GMGN blocked the request.")
                if not USE_CURL:
                    print("  RECOMMENDED FIX: install curl_cffi to bypass Cloudflare TLS checks:")
                    print("    pip install curl_cffi")
                    print("  Then re-run the script — no cookie needed with curl_cffi.\n")
                else:
                    print("  curl_cffi is installed but still 403.")
                    print("  Your cf_clearance cookie may have expired.")
                    print("  Grab a fresh one from Chrome DevTools -> Network -> any /rank/ request")
                    print("  and update COOKIE_STRING at the top of this script.\n")
                break

            resp.raise_for_status()
            data = resp.json()

        except requests.exceptions.HTTPError as e:
            print(f"  [ERROR] HTTP {e}")
            break
        except Exception as e:
            print(f"  [ERROR] {e}")
            break

        items = (
            data.get("data", {}).get("rank") or
            data.get("data", {}).get("tokens") or
            data.get("data") or
            data.get("tokens") or
            []
        )
        if isinstance(items, dict):
            items = list(items.values())

        if not items:
            print("  No more tokens returned. Stopping.")
            break

        tokens.extend(items)
        print(f"  Got {len(items)} tokens. Total: {len(tokens)}")

        if len(items) < limit:
            break

        offset += limit
        time.sleep(0.5)

    return tokens[:TOP_N]


def fmt_num(val):
    if val is None:
        return "N/A"
    try:
        f = float(val)
        if f >= 1_000_000:
            return f"${f/1_000_000:.2f}M"
        if f >= 1_000:
            return f"${f/1_000:.1f}K"
        return f"${f:.2f}"
    except (ValueError, TypeError):
        return str(val)


def print_tokens(tokens):
    print(f"\n{'='*60}")
    print(f" Results: {len(tokens)} tokens")
    print(f"{'='*60}")
    for t in tokens:
        print(
            f"\n  [{t.get('rank','?'):>3}]  {t.get('name','?')}  ({t.get('symbol','?')})"
            f"\n       Address    : {t.get('address','N/A')}"
            f"\n       Volume     : {fmt_num(t.get('volume'))}"
            f"\n       Liquidity  : {fmt_num(t.get('liquidity'))}"
            f"\n       Market Cap : {fmt_num(t.get('market_cap'))}"
            f"\n       Buy/Sell   : {t.get('buy_tax','?')}% / {t.get('sell_tax','?')}%"
            f"\n       Honeypot   : {bool(t.get('is_honeypot'))}  |  "
            f"Renounced: {bool(t.get('is_renounced'))}  |  "
            f"Rug: {t.get('rug_ratio','?')}  |  "
            f"Wash: {t.get('is_wash_trading','?')}"
        )
    print(f"\n{'='*60}\n")


def save_addresses(tokens, filename="base.txt"):
    addrs = [t["address"] for t in tokens if t.get("address")]
    with open(filename, "w") as f:
        f.write("\n".join(addrs))
    print(f"[OK] {len(addrs)} addresses saved -> {filename}")


def col_letter(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def save_excel(tokens, filename="gmgn_tokens.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{CHAIN.upper()} {TIMEFRAME}"

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="1A1A2E")
    alt_fill  = PatternFill("solid", start_color="F0F4FF")
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")
    thin      = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    num_cols = 17
    ws.merge_cells(f"A1:{col_letter(num_cols)}1")
    meta           = ws["A1"]
    meta.value     = (f"GMGN  |  {CHAIN.upper()}  |  {TIMEFRAME}  |  "
                      f"Fetched: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                      f"Top {len(tokens)} by Volume")
    meta.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    meta.fill      = PatternFill("solid", start_color="0F3460")
    meta.alignment = center
    ws.row_dimensions[1].height = 22

    headers = [
        "Rank","Name","Symbol","Address",
        "Volume ($)","Liquidity ($)","Market Cap ($)",
        "Buy Tax (%)","Sell Tax (%)",
        "Honeypot","Renounced",
        "Rug Ratio","Sniper Count","Bundler Rate",
        "Dev Hold %","Top70 Sniper %","Wash Trading",
    ]
    keys = [
        "rank","name","symbol","address",
        "volume","liquidity","market_cap",
        "buy_tax","sell_tax",
        "is_honeypot","is_renounced",
        "rug_ratio","sniper_count","bundler_rate",
        "dev_team_hold_rate","top70_sniper_hold_rate","is_wash_trading",
    ]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin
    ws.row_dimensions[2].height = 18

    for ri, t in enumerate(tokens, 3):
        fill = alt_fill if ri % 2 == 1 else None
        for ci, key in enumerate(keys, 1):
            val = t.get(key)
            if key in ("is_honeypot", "is_renounced", "is_wash_trading"):
                val = "Yes" if val else "No"
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.alignment = left if ci == 4 else center
            if fill:
                c.fill = fill
            if key in ("volume", "liquidity", "market_cap"):
                c.number_format = '#,##0.00'
            elif key in ("buy_tax","sell_tax","rug_ratio","bundler_rate",
                         "dev_team_hold_rate","top70_sniper_hold_rate"):
                c.number_format = '0.00'

    widths = [6, 20, 10, 46, 16, 16, 16, 10, 10, 10, 10, 10, 12, 12, 12, 14, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[col_letter(ci)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{col_letter(len(headers))}2"

    wb.save(filename)
    print(f"[OK] Excel saved -> {filename}")


def main():
    tokens = fetch_tokens()
    if not tokens:
        print("\n[!] No tokens fetched.\n")
        return
    print_tokens(tokens)
    save_addresses(tokens, "base.txt")
    save_excel(tokens, "gmgn_tokens.xlsx")
    print(f"\n[OK] Done — {len(tokens)} tokens processed.\n")


if __name__ == "__main__":
    main()
